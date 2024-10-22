# *******************************************************************************************************************
#   Energy Bill Analyser
# *******************************************************************************************************************
# Aim: Analyse the data from aurora for the last year of energy usage
# 
# *******************************************************************************************************************
# Version Control
# *******************************************************************************************************************
# V01   - Initial offering
#
#
# *******************************************************************************************************************
# Bug reports / ToDo
# *******************************************************************************************************************
# 
# 
# *******************************************************************************************************************
# Imports
# *******************************************************************************************************************

import os
import openpyxl
import numpy as np
import matplotlib.pyplot as plt
import collections
from datetime import datetime



# *******************************************************************************************************************
# Settings
# *******************************************************************************************************************

wb_fn = "C:\scratch\Your Energy Summary - 102293739\Your Energy Summary - 102293739.xlsx"



# *******************************************************************************************************************
# Functions
# *******************************************************************************************************************
class energy_date:
    def __init__(self, year, month, day, hour):
        self.year = year

# *******************************************************************************************************************
# Main
# *******************************************************************************************************************

# Load excel wb
wb = openpyxl.load_workbook(wb_fn)
ws = wb.active

print('Total number of rows: ' + str(ws.max_row) + '. And total number of columns: ' + str(ws.max_column))

tariff_31 = list()  # Residential light and power - Tariff 31
tariff_41 = list()  # Heating and Hot Water - Tariff 41

# Seperate tariff 31 and tariff 41 values
for value in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column, values_only=True):
    if value[3] == 'TAS31':
        tariff_31.append(value)
    elif value[3] == 'TAS41':
        tariff_41.append(value)
    else:
        # Add the header (first row) to both lists
        tariff_31.append(value)
        tariff_41.append(value)

tariff_31_hour = {}
tariff_41_hour = {}
dt_format_excel = "%m/%d/%Y %I:%M:%S %p"


# data import from excel reads values wrong. any datetime obj has .day and .month switched. Fixed in below for loops.
for value in tariff_31[1:]:     # tariff_31[1:] will skip first entry in tariff_31, the header row
    try:
        tariff_31_hour[str(value[2].year) + '-' + str(value[2].day) + '-' + str(value[2].month) + '-' + str(value[2].hour)] += value[6] # value[2].day is actually month
    except KeyError:
        # KeyError occured, make the key in energy_by_hour dict
        tariff_31_hour[str(value[2].year) + '-' + str(value[2].day) + '-' + str(value[2].month) + '-' + str(value[2].hour)] = value[6]  # value[2].day is actually month
    except AttributeError:
        # Some value[2] in the source data come back as datetime objects, some dont. This is for those that dont.
        dt_temp = datetime.strptime(value[2], dt_format_excel)
        # need the try except here too, to catch first time key is used
        try:
            tariff_31_hour[str(dt_temp.year) + '-' + str(dt_temp.month) + '-' + str(dt_temp.day) + '-' + str(dt_temp.hour)] += value[6]
        except KeyError:
            tariff_31_hour[str(dt_temp.year) + '-' + str(dt_temp.month) + '-' + str(dt_temp.day) + '-' + str(dt_temp.hour)] = value[6]




for value in tariff_41[1:]:     # tariff_41[1:] will skip first entry in tariff_31, the header row
    try:
        tariff_41_hour[str(value[2].year) + '-' + str(value[2].day) + '-' + str(value[2].month) + '-' + str(value[2].hour)] += value[6] # value[2].day is actually month
    except KeyError:
        # KeyError occured because key does not exist. Expected to happen on first run through, make the key in energy_by_hour dict
        tariff_41_hour[str(value[2].year) + '-' + str(value[2].day) + '-' + str(value[2].month) + '-' + str(value[2].hour)] = value[6]  # value[2].day is actually month
    except AttributeError:
        # Some value[2] in the source data come back as datetime objects, some dont. This is for those that dont.
        dt_temp = datetime.strptime(value[2], dt_format_excel)
        # need the try except here too, to catch first time key is used
        try:
            tariff_41_hour[str(dt_temp.year) + '-' + str(dt_temp.month) + '-' + str(dt_temp.day) + '-' + str(dt_temp.hour)] += value[6]
        except KeyError:
            tariff_41_hour[str(dt_temp.year) + '-' + str(dt_temp.month) + '-' + str(dt_temp.day) + '-' + str(dt_temp.hour)] = value[6]

# Create dict to sum tariff_31 and tariff_41
counter = collections.Counter()
for d in [tariff_31_hour, tariff_41_hour]:
    counter.update(d)

sum_tariffs_hour = dict(counter)


# plot the energy usage of each
# fig, ax = plt.subplots(1, 1, figsize=(15, 3))
# ax.plot(tariff_31_hour.keys(), tariff_31_hour.values(), color='b', label='Tariff 31 - Residential light and power')
# ax.plot(tariff_41_hour.keys(), tariff_41_hour.values(), color='r', label='Tariff 41 - Heating and Hot Water')
# ax.plot(tariff_41_hour.keys(), sum_tariffs_hour.values(), color='g', label='Total')
# fig.tight_layout()
# fig.autofmt_xdate()
# plt.xticks(visible=False)
# ax.get_xaxis().set_visible(False)
# for label in ax.get_xaxis().get_ticklabels()[::6]:  # plot ever 6th x-axis label
#     label.set_visible(True)
# plt.setp(ax.axes.get_xticklabels(), visible=False)
# plt.setp(ax.axes.get_xticklabels()[::5], visible=True)
# ax.set_xticks(ax.get_xticks()[::6])


# plt.xlabel('Time')
# plt.ylabel('energy (kWH)')
# plt.title('Energy Usage')

# plt.legend()
# plt.show()

dt_format = "%Y-%m-%d-%H"

# create month dicts
avg_hour_tariff_31 = {}
avg_hour_tariff_41 = {}

# Group data by month/day, with the goal to get an average day of energy usage, for every month.

for key, value in tariff_31_hour.items():   # this loop sums all hours, for a given month
    # similar try except as above, try to add the hour value to itself, or create the dict key if it hasnt been created before
    try:
        avg_hour_tariff_31[datetime(1, datetime.strptime(key, dt_format).month, 1).strftime("%b") + '_' + key.rpartition('-')[2]] += value
    except KeyError:
        avg_hour_tariff_31[datetime(1, datetime.strptime(key, dt_format).month, 1).strftime("%b") + '_' + key.rpartition('-')[2]] = value

# do the same for the other tariff
for key, value in tariff_41_hour.items():   # this loop sums all hours, for a given month
    # similar try except as above, try to add the hour value to itself, or create the dict key if it hasnt been created before
    try:
        avg_hour_tariff_41[datetime(1, datetime.strptime(key, dt_format).month, 1).strftime("%b") + '_' + key.rpartition('-')[2]] += value
    except KeyError:
        avg_hour_tariff_41[datetime(1, datetime.strptime(key, dt_format).month, 1).strftime("%b") + '_' + key.rpartition('-')[2]] = value

# loop through values in avg_hour, and divide by days in the month
for key, value in avg_hour_tariff_31.items():
    if key[0:3] in ['Apr', 'Jun', 'Sep', 'Nov']:
        avg_hour_tariff_31[key] = avg_hour_tariff_31[key] / 30
    elif key[0:3] in ['Jan', 'Mar', 'May', 'Jul', 'Aug', 'Oct', 'Dec']:
        avg_hour_tariff_31[key] = avg_hour_tariff_31[key] / 31
    else:
        avg_hour_tariff_31[key] = avg_hour_tariff_31[key] / 28
        
for key, value in avg_hour_tariff_41.items():
    if key[0:3] in ['Apr', 'Jun', 'Sep', 'Nov']:
        avg_hour_tariff_41[key] = avg_hour_tariff_41[key] / 30
    elif key[0:3] in ['Jan', 'Mar', 'May', 'Jul', 'Aug', 'Oct', 'Dec']:
        avg_hour_tariff_41[key] = avg_hour_tariff_41[key] / 31
    else:
        avg_hour_tariff_41[key] = avg_hour_tariff_41[key] / 28


# create plots of average daily usage, by month
xaxis = np.array([0, 23])     # array of hours, from 0-23

yaxis_tariff_31 = {}
yaxis_tariff_41 = {}
yaxis_total = {}

for key, value in avg_hour_tariff_31.items():
    try:
        yaxis_tariff_31[key[0:3]].append(value)
    except KeyError:
        yaxis_tariff_31[key[0:3]] = [value]

for key, value in avg_hour_tariff_41.items():
    try:
        yaxis_tariff_41[key[0:3]].append(value)
    except KeyError:
        yaxis_tariff_41[key[0:3]] = [value]

for key, value in yaxis_tariff_31.items():
    yaxis_total[key] = [x + y for x, y in zip(value, yaxis_tariff_41[key])]

fig, ax = plt.subplots(1, 1, figsize=(15, 3))
ax.plot(yaxis_tariff_31['Jan'], color='b', label='January 31')
ax.plot(yaxis_tariff_41['Jan'], color='r', label='January 41')
ax.plot(yaxis_total['Jan'], color='g', label='January Total')

plt.xlabel('Time (hour)')
plt.ylabel('energy (kWH)')
plt.title('Average Daily Energy Usage')

plt.legend()
plt.show()


# calculate costs
tariff_31_day = 113.772     # c/day supply charge
tariff_31_usage = 29.947    # c/kWh
tariff_41_day = 21.214      # c/day supply charge
tariff_41_usage = 19.447    # c/kWh

peak_usage = 36.198         # c/kWh
offpeak_usage = 16.855      # c/kWh
peak_day = 126.389          # c/day supply charge


peak_hours = [7, 8, 9, 16, 17, 18, 19, 20]      # 7am - 10am, 4pm - 9pm
offpeak_hours = [0, 1, 2, 3, 4, 5, 6, 10, 11, 12, 13, 14, 15, 21, 22, 23]

for key, value in yaxis_total.items():
    month_flat = (sum(yaxis_tariff_31[key]) * tariff_31_usage) + (sum(yaxis_tariff_41[key]) * tariff_41_usage) + tariff_31_day + tariff_41_day
    month_peak = peak_day
    for hour, value in enumerate(value):
        if hour in peak_hours:
            month_peak += (value * peak_usage)
        else:
            month_peak += (value * offpeak_usage)

    print(key + ' avg flat cost (c): ' + str(month_flat))
    print(key + ' avg peak cost (c): ' + str(month_peak))


print('end')